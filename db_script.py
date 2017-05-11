from openpyxl import Workbook, load_workbook
from pymongo import MongoClient

def readDoc(db_name, db_type):
    client = MongoClient('mongodb://localhost:27017/')
    db = client["clicksocial"]

    wb = load_workbook(db_name)
    ws = wb[wb.sheetnames[0]]
    x = list(ws.rows)
    # Read columns and rows
    cols_data = x[0]
    rows_data = x[1:]
    # Create dict
    bulk = db["organizations"].initialize_ordered_bulk_op()
    for row in rows_data:
        pre_dataset = {}
        for idx, r in enumerate(list(row)):
            if r.value is not None:
                t = type(r.value)
                if t == float:
                    r.value = int(r.value)
                elif t.__name__ == 'datetime':
                    r.value = r.value.strftime('%d/%m/%Y')

                pre_dataset[cols_data[idx].value] = r.value

        pre_dataset["type"] = db_type
        if len(pre_dataset) > 0:
            # Create bulk op (Upsert)
            selection = pre_dataset
            bulk.find(selection).upsert().replace_one(pre_dataset)
    # Create new collection
    try:
        bulk.execute()
    except BulkWriteError:
        print("Error")



if __name__ == "__main__":
    readDoc("investigacion.xlsx", "Investigación")
